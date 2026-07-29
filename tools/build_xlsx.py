# -*- coding: utf-8 -*-
"""בונה את קובץ האקסל של התפקידים מתוך tools/role_stats.json.

    python3 tools/build_xlsx.py

שני גיליונות:
  "תפקידים וסיכויים"  — שורה לכל תפקיד: סיכוי, מאפיינים, תנאי סף, מקור.
  "מקרא"              — הסבר על כל עמודה ועל מה שהמספרים כן ולא אומרים.
"""
import io
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "tools", "role_stats.json")
OUT = os.path.join(ROOT, "תפקידים-וסיכויים.xlsx")

NAVY = "0D1B3E"
NAVY_SOFT = "1E3A63"
ACCENT_SOFT = "DBEAFE"
LINE = "E2E8F0"
ZEBRA = "F8FAFC"

TRACK_FILL = {
    "ליבה": "DBEAFE",
    "התמחות": "FEF3C7",
    "מנהלה": "E9E5F8",
}

with io.open(SRC, encoding="utf-8") as fh:
    data = json.load(fh)

roles = data["roles"]
N = data["n"]

# ליבה תחילה, ובתוך כל מסלול לפי הסיכוי.
track_order = {"ליבה": 0, "התמחות": 1, "מנהלה": 2}
roles.sort(key=lambda r: (track_order.get(r["track"], 9), -r["pctTop3"]))

wb = Workbook()
ws = wb.active
ws.title = "תפקידים וסיכויים"
ws.sheet_view.rightToLeft = True

COLUMNS = [
    ("תפקיד", 30),
    ("מסלול", 12),
    ("סיווג", 14),
    ("סיכוי להופיע ב-3 המובילים", 13),
    ("סיכוי להיות מספר 1", 12),
    ("אחוז ההתאמה הממוצע כשהוא מופיע", 12),
    ("שיעור הפרופילים שנחסמו בתנאי סף", 12),
    ("מה מאפיין את התפקיד", 46),
    ("מה מוביל אליו בשאלון", 52),
    ("תנאי סף", 34),
    ("דרישות כפי שפורסמו", 46),
    ("שכר", 26),
    ("מקור המידע", 18),
]

# ── כותרת עליונה ───────────────────────────────────────────────────────────
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
title = ws.cell(row=1, column=1, value="מצפן הקריירה — התפקידים והסיכוי שכל אחד יומלץ")
title.font = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
title.fill = PatternFill("solid", fgColor=NAVY)
title.alignment = Alignment(horizontal="right", vertical="center")
ws.row_dimensions[1].height = 30

ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNS))
sub = ws.cell(
    row=2,
    column=1,
    value="הסיכוי נמדד בסימולציה על %s פרופילים אקראיים. זהו הסיכוי שהכלי יציע את התפקיד "
          "למועמד אקראי — ולא הסיכוי להתקבל אליו בפועל. ראו גיליון \"מקרא\"." % f"{N:,}",
)
sub.font = Font(name="Calibri", size=10, color="475569")
sub.fill = PatternFill("solid", fgColor=ZEBRA)
sub.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 30

# ── שורת הכותרות ───────────────────────────────────────────────────────────
HEADER_ROW = 3
thin = Side(style="thin", color=LINE)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for i, (name, width) in enumerate(COLUMNS, start=1):
    c = ws.cell(row=HEADER_ROW, column=i, value=name)
    c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
    ws.column_dimensions[get_column_letter(i)].width = width
ws.row_dimensions[HEADER_ROW].height = 46


def driver_text(role):
    """מה מוביל לתפקיד — מתוך מטריצת הניקוד עצמה."""
    return "\n".join(
        "• %s  (+%d)  ← %s" % (d["option"], d["pts"], d["q"])
        for d in role["drivers"]
    )


def characterises(role):
    bits = [role["oneLiner"]]
    if role.get("dayInLife"):
        bits.append("היום-יום: " + role["dayInLife"])
    return "\n\n".join(b for b in bits if b)


row = HEADER_ROW + 1
for idx, r in enumerate(roles):
    never = r["pctTop3"] == 0
    values = [
        r["name"],
        r["track"],
        r["tier"] or "—",
        r["pctTop3"] / 100.0,
        r["pctFirst"] / 100.0,
        (r["avgMatch"] / 100.0) if r["avgMatch"] is not None else None,
        r["pctBlocked"] / 100.0,
        characterises(r),
        driver_text(r),
        r["gates"],
        "\n".join("• " + x for x in r["requirements"]),
        r["salaryShown"],
        r["source"],
    ]
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = border
        c.font = Font(name="Calibri", size=10)
        wrap = i in (8, 9, 10, 11, 12)
        c.alignment = Alignment(
            horizontal="right" if i in (1, 8, 9, 10, 11, 12) else "center",
            vertical="top" if wrap else "center",
            wrap_text=wrap,
        )
        if i in (4, 5, 6, 7):
            c.number_format = "0.0%"
        if i == 1:
            c.font = Font(name="Calibri", size=10, bold=True)
        if idx % 2 == 1:
            c.fill = PatternFill("solid", fgColor=ZEBRA)
        if i == 2:
            c.fill = PatternFill("solid", fgColor=TRACK_FILL.get(r["track"], "FFFFFF"))
        if never:
            # תפקיד שלא ניתן לאשר מהשאלון — מסומן בבירור.
            c.font = Font(name="Calibri", size=10, color="B45309",
                          bold=(i == 1), italic=True)
            c.fill = PatternFill("solid", fgColor="FFFBEB")
    ws.row_dimensions[row].height = 96
    row += 1

LAST = row - 1

# פס נתונים על עמודת הסיכוי — קריא במבט אחד.
ws.conditional_formatting.add(
    "D%d:D%d" % (HEADER_ROW + 1, LAST),
    DataBarRule(start_type="num", start_value=0, end_type="num", end_value=0.6,
                color=ACCENT_SOFT, showValue=True),
)

ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=2)
ws.auto_filter.ref = "A%d:%s%d" % (HEADER_ROW, get_column_letter(len(COLUMNS)), LAST)

# ── גיליון מקרא ────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("מקרא")
ws2.sheet_view.rightToLeft = True
ws2.column_dimensions["A"].width = 34
ws2.column_dimensions["B"].width = 96

ws2.merge_cells("A1:B1")
t2 = ws2.cell(row=1, column=1, value="מקרא — איך לקרוא את הטבלה")
t2.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
t2.fill = PatternFill("solid", fgColor=NAVY)
t2.alignment = Alignment(horizontal="right", vertical="center")
ws2.row_dimensions[1].height = 28

LEGEND = [
    ("סיכוי להופיע ב-3 המובילים",
     "מתוך %s פרופילים אקראיים, באחוז הזה מהם התפקיד הופיע באחת משלוש ההמלצות. "
     "זהו הסיכוי שהכלי יציע את התפקיד — לא הסיכוי להתקבל אליו, ולא מספר המשרות הפתוחות." % f"{N:,}"),
    ("סיכוי להיות מספר 1",
     "באחוז הזה מהפרופילים התפקיד הגיע למקום הראשון."),
    ("אחוז ההתאמה הממוצע",
     "כשהתפקיד כן הופיע — מה היה אחוז ההתאמה הממוצע שלו. אחוז גבוה עם סיכוי נמוך "
     "מציין תפקיד ממוקד: הוא מתאים מאוד, אבל רק למי שסימן את הדברים הנכונים."),
    ("שיעור הפרופילים שנחסמו",
     "באחוז הזה מהפרופילים התפקיד נחסם בתנאי סף. 100% פירושו שההסמכה הנדרשת "
     "אינה נשאלת בשאלון כלל, ולכן הוא לעולם לא יומלץ."),
    ("מה מוביל אליו בשאלון",
     "התשובות שתורמות לתפקיד הכי הרבה נקודות, עם מספר הנקודות. זה לא ניסוח חופשי — "
     "זה נלקח ישירות ממטריצת הניקוד שבקוד."),
    ("מסלול",
     "ליבה = סייר, בלש, חוקר, משל\"ט, סייר תנועה — עיקר התקנים, וללא תנאי סף חוסמים. "
     "התמחות = יחידות מיוחדות, מז\"פ, חקירות ייעודיות. מנהלה = תפקידים תומכים."),
    ("סיווג",
     "ליבה ראשית = חוקר/בלש/סייר, שמקבלים משבצת שמורה. ליבה משנית = משל\"ט וסייר תנועה. "
     "מנהלה מרכזית = שער כניסה שנפתח בניסיון כללי. מנהלה מקצועית = דורש הסמכה ספציפית."),
    ("שורות בצהוב",
     "תפקידים שההסמכה שלהם אינה נשאלת בשאלון (תואר בתזונה, רישיון רפואי). הם מופיעים "
     "בקטלוג ובבלוק השקיפות עם הסיבה, אך לעולם אינם מומלצים — אין דרך לאשר אותם מהתשובות."),
    ("שכר",
     "\"נמסר ביום המיון\" פירושו שהמקור אינו מפרסם את הנתון. לא הושלם כאן שום טווח משוער."),
    ("מקור המידע",
     "מאגר מרכז הגיוס = שכר, הכשרה והתקדמות מפורסמים. דף התפקידים הרשמי = תיאור ודרישות "
     "בלבד; שכר והכשרה אינם מתפרסמים שם ונמסרים במרכז הגיוס."),
]

r2 = 3
for name, text in LEGEND:
    a = ws2.cell(row=r2, column=1, value=name)
    a.font = Font(name="Calibri", size=10, bold=True)
    a.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
    a.fill = PatternFill("solid", fgColor=ACCENT_SOFT)
    a.border = border
    b = ws2.cell(row=r2, column=2, value=text)
    b.font = Font(name="Calibri", size=10)
    b.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
    b.border = border
    ws2.row_dimensions[r2].height = 44
    r2 += 1

wb.save(OUT)
print("נכתב:", OUT)
print("תפקידים:", len(roles), "| פרופילים בסימולציה:", N)
